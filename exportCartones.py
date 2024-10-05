import json
import os
import random
from uuid import uuid4
from openpyxl import Workbook
from openpyxl.styles import Alignment
import sys
import ctypes

# Modelos predefinidos: tu lista de posiciones fijas
modelos = [
    [
        (0, 0), (2, 0), (0, 1), (1, 1), (2, 2), (1, 3), (0, 4), (1, 4),
        (2, 5), (1, 6), (2, 6), (0, 7), (1, 7), (0, 8), (2, 8)
    ],  # Modelo 1
    [
        (0, 0), (1, 0), (0, 1), (2, 1), (1, 2), (2, 2), (0, 3), (1, 3),
        (2, 4), (0, 5), (2, 5),  (0, 6), (1, 6), (2, 7), (1, 8)
    ],  # Modelo 2
    [
        (2, 0), (0, 1), (0, 2), (1, 2), (2, 3), (1, 4), (2, 4), (0, 5),
        (1, 5), (0, 6), (2, 6), (1, 7), (2, 7), (0, 8), (1, 8)
    ],  # Modelo 3
    [
        (1, 0), (0, 1), (2, 1), (0, 2), (1, 2), (1, 3), (2, 3), (2, 4),
        (1, 5), (2, 5), (2, 6), (0, 7), (1, 7), (0, 8), (2, 8)
    ]  # Modelo 4
]

def marcar_como_oculto(archivo):
    try:
        ctypes.windll.kernel32.SetFileAttributesW(archivo, 2)  # 2 = FILE_ATTRIBUTE_HIDDEN
    except AttributeError:
        pass  # Si no estás en Windows, esta llamada no hará nada

# Crear directorio si no existe
def crear_directorio_si_no_existe(directorio):
    os.makedirs(directorio, exist_ok=True)

# Generación de un único cartón de bingo con idCarton y el modelo seleccionado
def generar_carton(id_carton, modelo):
    carton = [[None for _ in range(9)] for _ in range(3)]
    rango_limites = [(1, 10), (11, 20), (21, 30), (31, 40), (41, 50), (51, 60), (61, 70), (71, 80), (81, 90)]
    usados = set()

    # Colocar números en las posiciones del modelo
    for fila, columna in modelos[modelo]:
        numero = random.randint(*rango_limites[columna])
        while numero in usados:
            numero = random.randint(*rango_limites[columna])
        carton[fila][columna] = numero
        usados.add(numero)

    return {"idCarton": id_carton, "carton": carton}

# Función que genera cartones por bloques y distribuidos en 4 modelos
def generar_cartones_en_bloques(cantidad, bloque_size, serial, output_dir, jugada_numero, fecha):
    todos_los_cartones = []
    todos_los_cartones_final = []  # Lista que acumula todos los cartones para el JSON final
    archivos_json = []
    bloque_numero = 1

    for i in range(cantidad):
        modelo = i % 4  # Rotamos entre los 4 modelos
        carton = generar_carton(i + 1, modelo)  # El id del cartón será i+1
        carton['jugada'] = jugada_numero  # Agregar el número de jugada al cartón
        carton['fecha'] = fecha  # Agregar la fecha al cartón
        todos_los_cartones.append(carton)
        todos_los_cartones_final.append(carton)  # Acumular todos los cartones

        # Guardar cartones en bloques cuando se alcanza el tamaño de bloque
        if (i + 1) % bloque_size == 0 or (i + 1) == cantidad:
            archivo_json = os.path.join(output_dir, f"{serial}_cartones_bloque_{bloque_numero}.json")
            with open(archivo_json, 'w', encoding='utf-8') as f:
                json.dump(todos_los_cartones, f, indent=2)
            archivos_json.append(archivo_json)
            
            # Marcar el archivo como oculto
            marcar_como_oculto(archivo_json)
            
            print(f"Guardado bloque {bloque_numero} en {archivo_json}")
            todos_los_cartones.clear()  # Limpiar la lista para el próximo bloque
            bloque_numero += 1

    return archivos_json, todos_los_cartones_final  # Retorna los archivos JSON y todos los cartones acumulados

# Exportar a un archivo Excel con una hoja por cada modelo
def exportar_a_excel_con_hojas(archivo_excel, serial, archivos_json):
    workbook = Workbook()

    # Crear una hoja por cada modelo
    hojas_modelo = {}
    for modelo in range(4):
        sheet = workbook.create_sheet(f"Modelo {modelo + 1}")
        sheet.append(["ID", "Fecha", "Nº Jugada", "Serial", "C1 F1", "C1 F2", "C1 F3", "C2 F1", "C2 F2", "C2 F3", "C3 F1", "C3 F2", "C3 F3",
                      "C4 F1", "C4 F2", "C4 F3", "C5 F1", "C5 F2", "C5 F3", "C6 F1", "C6 F2", "C6 F3", "C7 F1", 
                      "C7 F2", "C7 F3", "C8 F1", "C8 F2", "C8 F3", "C9 F1", "C9 F2", "C9 F3"])
        hojas_modelo[modelo] = sheet

    # Leer cada archivo JSON y distribuir los cartones en las hojas correspondientes a su modelo
    for archivo_json in archivos_json:
        # Cargar el contenido del archivo JSON
        with open(archivo_json, 'r', encoding='utf-8') as f:
            cartones = json.load(f)  # Aquí cargamos correctamente el JSON como lista de diccionarios

            for carton_obj in cartones:
                modelo = carton_obj["idCarton"] % 4  # Determinar el modelo del cartón (asignado ciclicamente)
                carton = carton_obj["carton"]
                id_carton = carton_obj["idCarton"]
                jugada = carton_obj["jugada"]
                fecha = carton_obj["fecha"]
                fila_carton = [id_carton, fecha, jugada, serial]  # Primera columna ID, segunda Fecha, tercera jugada, cuarta el serial
                for col_index in range(9):
                    fila_carton.append(carton[0][col_index] if carton[0][col_index] is not None else '-')
                    fila_carton.append(carton[1][col_index] if carton[1][col_index] is not None else '-')
                    fila_carton.append(carton[2][col_index] if carton[2][col_index] is not None else '-')
                hojas_modelo[modelo].append(fila_carton)

    # Remover la hoja por defecto si existe
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    # Guardar el archivo Excel
    workbook.save(archivo_excel)
    print(f"Archivo Excel final guardado en {archivo_excel}")

# Exportar todos los cartones a un archivo JSON final
def exportar_cartones_final(archivo_final_json, todos_los_cartones):
    # Verificar cuántos cartones se van a exportar
    print(f"Total de cartones a exportar en el JSON final: {len(todos_los_cartones)}")
    
    # Asegúrate de que hay cartones en la lista
    if todos_los_cartones:
        with open(archivo_final_json, 'w', encoding='utf-8') as f:
            json.dump(todos_los_cartones, f, indent=2)
        print(f"Archivo final de cartones guardado en {archivo_final_json}")
    else:
        print("No se encontraron cartones para exportar al JSON final.")

# Exportar serial a un archivo txt
def exportar_serial_a_txt(serial, fecha, jugada_numero, output_dir):
    archivo_txt = os.path.join(output_dir, f"{serial}_serial.txt")
    with open(archivo_txt, 'w') as file:
        file.write(f"Serial: {serial}\nN jugada: {jugada_numero}\nFecha: {fecha}")
    print(f"Serial exportado exitosamente a {archivo_txt}")

if __name__ == "__main__":
    # Leer argumentos
    cantidad = int(sys.argv[1])  # Cantidad total de cartones
    bloque_size = int(sys.argv[2])  # Tamaño de los bloques de guardado
    archivo_excel = sys.argv[3]  # Ruta del archivo Excel
    jugada_numero = sys.argv[4]  # Número de jugada
    fecha_juego = sys.argv[5]  # Fecha del juego

    # Crear un serial único
    serial = str(uuid4())

    # Crear carpetas necesarias
    serial_root = os.path.join(os.path.dirname(archivo_excel), 'seriales')
    crear_directorio_si_no_existe(serial_root)
    output_dir = os.path.join(serial_root, serial)
    crear_directorio_si_no_existe(output_dir)

    # Actualizar la ruta del archivo Excel para que esté dentro de la carpeta del serial
    archivo_excel = os.path.join(output_dir, f"{serial}_{fecha_juego}_cartones.xlsx")
    archivo_final_json = os.path.join(output_dir, f"{serial}_{fecha_juego}_cartones_final.json")

    # Generar cartones distribuidos en los 4 modelos y guardarlos por bloques
    print(f"Generando {cantidad} cartones en bloques de {bloque_size}...")
    archivos_json, todos_los_cartones = generar_cartones_en_bloques(cantidad, bloque_size, serial, output_dir, jugada_numero, fecha_juego)

    # Exportar serial a archivo txt
    exportar_serial_a_txt(serial, fecha_juego, jugada_numero, output_dir)

    # Exportar a Excel
    print(f"Exportando {cantidad} cartones a Excel...")
    exportar_a_excel_con_hojas(archivo_excel, serial, archivos_json)

    # Exportar todos los cartones a un archivo JSON final
    print(f"Exportando todos los cartones a un archivo JSON final...")
    exportar_cartones_final(archivo_final_json, todos_los_cartones)

    print(f"Proceso completado. Archivos generados en {output_dir}")
